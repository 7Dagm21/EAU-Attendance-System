import csv
from io import BytesIO
from datetime import date, timedelta
from django.http import HttpResponse
from django.db.models import Sum, Avg
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer
)
from reportlab.graphics.shapes import Drawing, Rect, String, Line, Wedge, Circle
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics import renderPDF
from reportlab.platypus import Image as RLImage
from decimal import Decimal
from .models import Student, AttendanceRecord, Course, CourseOffering, Enrollment
from decimal import Decimal as _D
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

LATE_DEDUCTION = _D('0.5')
ACTIVATION_THRESHOLD = _D('0.15')


def _short_date(d):
    """Format a date (or ISO date string) as 'M/D' for compact display."""
    if isinstance(d, str):
        d = date.fromisoformat(d)
    return f"{d.month}/{d.day}"


# ─────────────────────────────────────────
# COLORS
# ─────────────────────────────────────────
DARK_GREEN  = colors.HexColor('#1A4A0F')
MID_GREEN   = colors.HexColor('#2D6A1F')
LIGHT_GREEN = colors.HexColor('#E8F5E0')
RED         = colors.HexColor('#C0392B')
ORANGE      = colors.HexColor('#E67E22')
GREEN       = colors.HexColor('#27AE60')
GRAY        = colors.HexColor('#555555')
LIGHT_GRAY  = colors.HexColor('#F5F5F5')
WHITE       = colors.white
BLUE        = colors.HexColor('#2563EB')
AMBER       = colors.HexColor('#D97706')

# XLSX hex (no #)
XL_HEADER_BG   = '1A4A0F'
XL_HEADER_FONT = 'FFFFFF'
XL_SAFE_BG     = 'D1FAE5'
XL_WARN_BG     = 'FEF3C7'
XL_RISK_BG     = 'FEE2E2'
XL_SAFE_FG     = '065F46'
XL_WARN_FG     = '92400E'
XL_RISK_FG     = '991B1B'
XL_ALT_ROW     = 'F0FDF4'
XL_SUBHEADER   = '2D6A1F'
XL_SECTION_BG  = 'EFF6FF'
XL_SECTION_FG  = '1E40AF'


# ─────────────────────────────────────────
# CALC HELPERS
# ─────────────────────────────────────────
def calc_attendance(present_hours, late_hours, total_credit_hours, late_session_count=1):
    total_penalty = LATE_DEDUCTION * _D(str(late_session_count))
    earned_late = max(late_hours - total_penalty, _D('0'))
    earned = present_hours + earned_late
    pct = round(float(earned / total_credit_hours * 100) if total_credit_hours > 0 else 0.0, 1)
    return earned, pct


def is_gate_open(total_logged_hours, total_credit_hours):
    if total_credit_hours <= 0:
        return False
    return total_logged_hours >= total_credit_hours * ACTIVATION_THRESHOLD


def get_status_color_rl(status):
    if status == 'Safe':
        return GREEN
    elif status == 'Warning':
        return ORANGE
    return RED


def format_date_range(filter_meta):
    report_type = filter_meta.get('report_type', 'full')
    start = filter_meta.get('start_date')
    end = filter_meta.get('end_date')
    if report_type == 'weekly' and start and end:
        try:
            from datetime import datetime
            s = datetime.strptime(str(start), '%Y-%m-%d').strftime('%d %b %Y')
            e = datetime.strptime(str(end),   '%Y-%m-%d').strftime('%d %b %Y')
            return f"Week: {s} – {e}", "Weekly"
        except Exception:
            pass
    if report_type == 'custom' and start and end:
        try:
            from datetime import datetime
            s = datetime.strptime(str(start), '%Y-%m-%d').strftime('%d %b %Y')
            e = datetime.strptime(str(end),   '%Y-%m-%d').strftime('%d %b %Y')
            return f"Custom Range: {s} – {e}", "Custom"
        except Exception:
            pass
    return "Full Semester", "Full Semester"


# ─────────────────────────────────────────
# LEGACY HELPERS (used by older paths)
# ─────────────────────────────────────────
def get_course_summary(course, start_date=None, end_date=None):
    offering_ids = CourseOffering.objects.filter(course=course).values_list('id', flat=True)
    student_ids = AttendanceRecord.objects.filter(
        course_offering_id__in=offering_ids
    ).values_list('student_id', flat=True).distinct()
    students = Student.objects.filter(id__in=student_ids)
    summary = []
    for student in students:
        attended_filters = {
            'student': student, 'course_offering__course': course, 'status__in': ['present', 'late'],
        }
        if start_date: attended_filters['date__gte'] = start_date
        if end_date:   attended_filters['date__lte'] = end_date
        attended_hours = AttendanceRecord.objects.filter(**attended_filters).aggregate(
            total=Sum('hours_attended'))['total'] or Decimal('0')
        missed_filters = {
            'student': student, 'course_offering__course': course, 'status__in': ['absent', 'excused'],
        }
        if start_date: missed_filters['date__gte'] = start_date
        if end_date:   missed_filters['date__lte'] = end_date
        missed_hours = AttendanceRecord.objects.filter(**missed_filters).aggregate(
            total=Sum('hours_attended'))['total'] or Decimal('0')
        total_hours = course.total_credit_hours
        minimum = Decimal(str(course.minimum_required_hours))
        total_recorded = attended_hours + missed_hours
        percentage = round(float(attended_hours / total_recorded * 100) if total_recorded > 0 else 100.0, 1)
        if percentage >= 90:   status = 'Safe'
        elif percentage >= 85: status = 'Warning'
        else:                  status = 'At Risk'
        summary.append({
            'student_id': student.student_id, 'full_name': student.full_name,
            'attended_hours': float(attended_hours), 'missed_hours': float(missed_hours),
            'total_hours': float(total_hours), 'percentage': percentage,
            'minimum_required': float(minimum), 'status': status
        })
    return summary


# ─────────────────────────────────────────
# PDF HEADER
# ─────────────────────────────────────────
def build_pdf_header(elements, title, subtitle, styles, filter_label=None, report_period=None):
    elements.append(Paragraph(
        "Ethiopian Aviation University",
        ParagraphStyle('uni', fontSize=10, textColor=GRAY, alignment=1, spaceAfter=3)
    ))
    elements.append(Paragraph(
        "Student Attendance Management System",
        ParagraphStyle('sys', fontSize=8, textColor=GRAY, alignment=1, spaceAfter=10)
    ))
    elements.append(Paragraph(
        title,
        ParagraphStyle('title', fontSize=18, textColor=DARK_GREEN, alignment=1,
                       spaceAfter=5, fontName='Helvetica-Bold')
    ))
    elements.append(Paragraph(
        subtitle,
        ParagraphStyle('sub', fontSize=10, textColor=MID_GREEN, alignment=1, spaceAfter=4)
    ))
    if report_period:
        elements.append(Paragraph(
            f"Period: {report_period}",
            ParagraphStyle('period', fontSize=9, textColor=BLUE, alignment=1,
                           spaceAfter=3, fontName='Helvetica-Bold')
        ))
    if filter_label:
        elements.append(Paragraph(
            f"Filter: {filter_label}",
            ParagraphStyle('filter', fontSize=9, textColor=GRAY, alignment=1, spaceAfter=3)
        ))
    elements.append(Paragraph(
        f"Generated: {date.today().strftime('%d %B %Y')}",
        ParagraphStyle('gendate', fontSize=9, textColor=GRAY, alignment=1, spaceAfter=14)
    ))
    elements.append(Spacer(1, 0.15 * inch))


# ─────────────────────────────────────────
# PDF BAR CHART (ReportLab native)
# ─────────────────────────────────────────
def build_pdf_bar_chart(band_data, width=300, height=185):
    """band_data: list of (label, count)"""
    if not band_data:
        return None
    labels = [b[0] for b in band_data]
    values = [b[1] for b in band_data]

    drawing = Drawing(width, height)

    chart = VerticalBarChart()
    chart.x = 50
    chart.y = 45
    chart.width = width - 70
    chart.height = height - 75
    chart.data = [values]
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.angle = 0
    chart.categoryAxis.labels.fontSize = 8
    chart.valueAxis.labels.fontSize = 8
    chart.valueAxis.forceZero = 1
    chart.bars[0].fillColor = BLUE
    chart.bars[0].strokeColor = None

    title_str = String(width / 2, height - 10, "Attendance Bands",
                       fontSize=9, fontName='Helvetica-Bold',
                       fillColor=DARK_GREEN, textAnchor='middle')
    # Legend: band explanations below the chart
    legend_text = String(width / 2, 30,
                         "Bands: <75% = At Risk  |  75–84.9% = Warning  |  85–89.9% = Near-Safe  |  ≥90% = Safe",
                         fontSize=6.5, fontName='Helvetica', fillColor=GRAY, textAnchor='middle')
    y_axis_label = String(10, height / 2, "No. of Students",
                          fontSize=7, fontName='Helvetica', fillColor=GRAY, textAnchor='middle')
    drawing.add(chart)
    drawing.add(title_str)
    drawing.add(legend_text)
    drawing.add(y_axis_label)
    return drawing


def build_pdf_pie_chart(risk_data, width=240, height=160):
    """risk_data: list of (label, value)"""
    if not risk_data:
        return None
    pie_colors = [GREEN, AMBER, RED]
    drawing = Drawing(width, height)

    pie = Pie()
    pie.x = 55
    pie.y = 25
    pie.width = 110
    pie.height = 110
    pie.data = [max(r[1], 0) for r in risk_data]
    pie.labels = [f"{r[0]}\n({r[1]})" for r in risk_data]
    pie.sideLabels = True
    pie.simpleLabels = False
    pie.slices.fontSize = 7
    for i, c in enumerate(pie_colors[:len(risk_data)]):
        pie.slices[i].fillColor = c
        pie.slices[i].strokeColor = WHITE
        pie.slices[i].strokeWidth = 1

    title_str = String(width / 2, height - 10, "Risk Distribution",
                       fontSize=9, fontName='Helvetica-Bold',
                       fillColor=DARK_GREEN, textAnchor='middle')
    drawing.add(pie)
    drawing.add(title_str)
    return drawing


# ─────────────────────────────────────────
# PDF — COURSE OFFERING REPORT
# ─────────────────────────────────────────
def generate_course_pdf(course, summary, title, filter_meta=None, offering=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    elements = []

    # Sort alphabetically
    summary = sorted(summary, key=lambda r: r['full_name'].lower())

    filter_meta = filter_meta or {}
    period_label, period_type = format_date_range(filter_meta)
    student_label = filter_meta.get('student_label')

    # Section/offering subtitle
    if offering:
        subtitle = (f"{course.name}  |  Section: {offering.section.name} "
                    f"(Year {offering.section.year})  |  "
                    f"Credit Hrs: {course.total_credit_hours}  |  "
                    f"Min Required: {course.minimum_required_hours} hrs")
    else:
        subtitle = (f"{course.name}  |  "
                    f"Credit Hrs: {course.total_credit_hours}  |  "
                    f"Min Required: {course.minimum_required_hours} hrs")

    build_pdf_header(
        elements, title, subtitle, styles,
        filter_label=student_label,
        report_period=period_label,
    )

    # ── KPI summary row ──────────────────
    safe    = sum(1 for r in summary if r['status'] == 'Safe')
    warning = sum(1 for r in summary if r['status'] == 'Warning')
    at_risk = sum(1 for r in summary if r['status'] == 'At Risk')
    total   = len(summary)
    avg_pct = round(sum(r['percentage'] for r in summary) / total, 1) if total else 0

    kpi_data = [[
        f"Total Students: {total}",
        f"Average Attendance: {avg_pct}%",
        f"Safe: {safe}",
        f"Warning: {warning}",
        f"At Risk: {at_risk}",
    ]]
    kpi_table = Table(kpi_data, colWidths=[140, 140, 100, 100, 100])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), LIGHT_GREEN),
        ('TEXTCOLOR', (0,0), (-1,-1), DARK_GREEN),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.2 * inch))

    # ── Charts ───────────────────────────
    band_data = []
    bands_raw = {'<75%': 0, '75–84.9%': 0, '85–89.9%': 0, '≥90%': 0}
    for row in summary:
        p = row['percentage']
        if p < 75:    bands_raw['<75%'] += 1
        elif p < 85:  bands_raw['75–84.9%'] += 1
        elif p < 90:  bands_raw['85–89.9%'] += 1
        else:         bands_raw['≥90%'] += 1
    band_data = list(bands_raw.items())
    risk_data = [('Safe', safe), ('Warning', warning), ('At Risk', at_risk)]

    bar_d  = build_pdf_bar_chart(band_data, width=310, height=185)
    pie_d  = build_pdf_pie_chart(risk_data, width=260, height=185)
    if bar_d and pie_d:
        chart_row = Table([[bar_d, pie_d]], colWidths=[340, 280])
        chart_row.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
        elements.append(chart_row)
        elements.append(Spacer(1, 0.2 * inch))

    # ── Student count row + main table ───
    count_data = [[f"Number of Students in this Report: {total}"]]
    count_table = Table(count_data, colWidths=[620])
    count_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), DARK_GREEN),
        ('TEXTCOLOR', (0,0), (-1,-1), WHITE),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
    ]))
    elements.append(count_table)

    table_data = [[
        '#', 'Student ID', 'Full Name', 'Present Hrs',
        'Late Hrs', 'Excused Hrs', 'Absent Hrs',
        'Attended %', 'Min Req.', 'Status', 'Excused Reason', 'Late/Absent Dates'
    ]]
    small_style = ParagraphStyle('small', fontSize=7, leading=9)
    for idx, row in enumerate(summary, start=1):
        late_d = ','.join(_short_date(d) for d in row.get('late_dates', []))
        absent_d = ','.join(_short_date(d) for d in row.get('absent_dates', []))
        dates_str = []
        if late_d:
            dates_str.append(f"L: {late_d}")
        if absent_d:
            dates_str.append(f"A: {absent_d}")
        table_data.append([
            str(idx),
            row['student_id'],
            row['full_name'],
            f"{row.get('present_hours', row.get('attended_hours', 0)):.1f}",
            f"{row.get('late_hours', 0):.1f}",
            f"{row.get('excused_hours', 0):.1f}",
            f"{row.get('absent_hours', row.get('missed_hours', 0)):.1f}",
            f"{row['percentage']}%",
            f"{row['minimum_required']:.0f} hrs",
            row['status'],
            row.get('excused_reason', ''),
            Paragraph('<br/>'.join(dates_str) or '—', small_style),
        ])

    col_widths = [22, 65, 115, 48, 42, 48, 48, 50, 45, 45, 90, 95]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), DARK_GREEN),
        ('TEXTCOLOR', (0,0), (-1,0), WHITE),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('ALIGN', (0,1), (-1,-1), 'CENTER'),
        ('ALIGN', (2,1), (2,-1), 'LEFT'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [WHITE, LIGHT_GREEN]),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
    ])
    for i, row in enumerate(summary, start=1):
        sc = get_status_color_rl(row['status'])
        style.add('TEXTCOLOR',  (9, i), (9, i), sc)
        style.add('FONTNAME',   (9, i), (9, i), 'Helvetica-Bold')
    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ─────────────────────────────────────────
# PDF — STUDENT REPORT
# ─────────────────────────────────────────
def generate_student_pdf(student, course_summaries, filter_meta=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=50, leftMargin=50,
                            topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    elements = []

    filter_meta = filter_meta or {}
    period_label, _ = format_date_range(filter_meta)

    build_pdf_header(
        elements,
        "Student Attendance Report",
        f"{student.full_name}  |  ID: {student.student_id}",
        styles,
        report_period=period_label,
    )

    total = len(course_summaries)
    avg_pct = round(sum(c['percentage'] for c in course_summaries) / total, 1) if total else 0
    safe    = sum(1 for c in course_summaries if c['status'] == 'Safe')
    warning = sum(1 for c in course_summaries if c['status'] == 'Warning')
    at_risk = sum(1 for c in course_summaries if c['status'] == 'At Risk')

    # KPI row
    kpi_data = [[
        f"Courses: {total}",
        f"Overall Avg: {avg_pct}%",
        f"Safe: {safe}",
        f"Warning: {warning}",
        f"At Risk: {at_risk}",
    ]]
    kpi_table = Table(kpi_data, colWidths=[90, 90, 70, 70, 70])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), LIGHT_GREEN),
        ('TEXTCOLOR', (0,0), (-1,-1), DARK_GREEN),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.15 * inch))

    # Pie chart
    risk_data = [('Safe', safe), ('Warning', warning), ('At Risk', at_risk)]
    pie_d = build_pdf_pie_chart(risk_data, width=250, height=160)
    if pie_d:
        elements.append(pie_d)
        elements.append(Spacer(1, 0.1 * inch))

    # Count row
    count_data = [[f"Number of Courses in this Report: {total}"]]
    count_table = Table(count_data, colWidths=[390])
    count_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), DARK_GREEN),
        ('TEXTCOLOR', (0,0), (-1,-1), WHITE),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(count_table)

    table_data = [['#', 'Course', 'Attended Hrs', 'Missed Hrs', 'Attendance %', 'Min Required', 'Status']]
    for idx, cs in enumerate(course_summaries, start=1):
        table_data.append([
            str(idx),
            cs['course_name'],
            f"{cs['attended_hours']:.1f} hrs",
            f"{cs['missed_hours']:.1f} hrs",
            f"{cs['percentage']}%",
            f"{cs['minimum_required']:.0f} hrs",
            cs['status'],
        ])

    col_widths = [28, 150, 75, 75, 75, 85, 70]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), DARK_GREEN),
        ('TEXTCOLOR', (0,0), (-1,0), WHITE),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('ALIGN', (0,1), (-1,-1), 'CENTER'),
        ('ALIGN', (1,1), (1,-1), 'LEFT'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [WHITE, LIGHT_GREEN]),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
    ])
    for i, cs in enumerate(course_summaries, start=1):
        c = get_status_color_rl(cs['status'])
        style.add('TEXTCOLOR', (6, i), (6, i), c)
        style.add('FONTNAME',  (6, i), (6, i), 'Helvetica-Bold')
    table.setStyle(style)
    elements.append(table)

    # Day-by-day late/absent breakdown
    day_rows = []
    for cs in course_summaries:
        for d in cs.get('late_dates', []):
            day_rows.append((d, cs['course_name'], 'Late'))
        for d in cs.get('absent_dates', []):
            day_rows.append((d, cs['course_name'], 'Absent'))
    day_rows.sort(key=lambda r: r[0])

    if day_rows:
        elements.append(Spacer(1, 0.2 * inch))
        heading_data = [["Days Marked Late or Absent"]]
        heading_table = Table(heading_data, colWidths=[390])
        heading_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), DARK_GREEN),
            ('TEXTCOLOR', (0,0), (-1,-1), WHITE),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(heading_table)

        day_table_data = [['#', 'Date', 'Course', 'Status']]
        from datetime import datetime as _dt
        for idx, (d, course_name, day_status) in enumerate(day_rows, start=1):
            try:
                d_display = _dt.strptime(str(d), '%Y-%m-%d').strftime('%a, %d %b %Y')
            except ValueError:
                d_display = str(d)
            day_table_data.append([str(idx), d_display, course_name, day_status])

        day_table = Table(day_table_data, colWidths=[28, 130, 150, 70], repeatRows=1)
        day_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), DARK_GREEN),
            ('TEXTCOLOR', (0,0), (-1,0), WHITE),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('ALIGN', (0,1), (-1,-1), 'CENTER'),
            ('ALIGN', (2,1), (2,-1), 'LEFT'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [WHITE, LIGHT_GREEN]),
            ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#CCCCCC')),
            ('TOPPADDING', (0,1), (-1,-1), 6),
            ('BOTTOMPADDING', (0,1), (-1,-1), 6),
        ])
        for idx, (_, _, day_status) in enumerate(day_rows, start=1):
            color = colors.HexColor('#D97706') if day_status == 'Late' else colors.HexColor('#DC2626')
            day_style.add('TEXTCOLOR', (3, idx), (3, idx), color)
            day_style.add('FONTNAME', (3, idx), (3, idx), 'Helvetica-Bold')
        day_table.setStyle(day_style)
        elements.append(day_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ─────────────────────────────────────────
# XLSX HELPERS
# ─────────────────────────────────────────
def _xl_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def _xl_font(bold=False, color='000000', size=10):
    return Font(bold=bold, color=color, size=size, name='Calibri')

def _xl_border():
    side = Side(style='thin', color='CCCCCC')
    return Border(left=side, right=side, top=side, bottom=side)

def _xl_center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _xl_left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def _xl_write_header_block(ws, row_start, info_rows):
    """Write a header info block. info_rows: list of (label, value) tuples."""
    for i, (label, value) in enumerate(info_rows):
        r = row_start + i
        ws.cell(r, 1).value = label
        ws.cell(r, 1).font = _xl_font(bold=True, color=XL_SUBHEADER)
        ws.cell(r, 1).alignment = _xl_left()
        ws.cell(r, 2).value = value
        ws.cell(r, 2).font = _xl_font(color='1F2937')
        ws.cell(r, 2).alignment = _xl_left()
    return row_start + len(info_rows)

def _xl_write_section_title(ws, row, text, col_span=10):
    ws.cell(row, 1).value = text
    ws.cell(row, 1).font = _xl_font(bold=True, color=XL_SECTION_FG, size=11)
    ws.cell(row, 1).fill = _xl_fill(XL_SECTION_BG)
    ws.cell(row, 1).alignment = _xl_left()
    if col_span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    return row + 1

def _xl_apply_table_header(ws, row, headers, col_start=1):
    for j, h in enumerate(headers):
        c = ws.cell(row, col_start + j)
        c.value = h
        c.fill = _xl_fill(XL_HEADER_BG)
        c.font = _xl_font(bold=True, color=XL_HEADER_FONT, size=10)
        c.alignment = _xl_center()
        c.border = _xl_border()

def _status_fill_font(status):
    if status == 'Safe':
        return _xl_fill(XL_SAFE_BG), _xl_font(bold=True, color=XL_SAFE_FG)
    elif status == 'Warning':
        return _xl_fill(XL_WARN_BG), _xl_font(bold=True, color=XL_WARN_FG)
    return _xl_fill(XL_RISK_BG), _xl_font(bold=True, color=XL_RISK_FG)

def _xl_add_bar_chart(ws, chart_data_row_start, num_categories, chart_anchor,
                       title="Attendance Bands", series_col=2, cat_col=1):
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.style = 10
    chart.y_axis.title = "Students"
    chart.x_axis.title = "Band"
    chart.width = 14
    chart.height = 9
    data_ref = Reference(ws, min_col=series_col, min_row=chart_data_row_start,
                         max_row=chart_data_row_start + num_categories - 1)
    cats_ref = Reference(ws, min_col=cat_col, min_row=chart_data_row_start,
                         max_row=chart_data_row_start + num_categories - 1)
    chart.add_data(data_ref)
    chart.set_categories(cats_ref)
    chart.series[0].title = None
    chart.series[0].graphicalProperties.solidFill = "2563EB"
    ws.add_chart(chart, chart_anchor)

def _xl_add_pie_chart(ws, chart_data_row_start, num_slices, chart_anchor,
                       title="Risk Distribution", val_col=2, cat_col=1):
    chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.width = 14
    chart.height = 9
    data_ref = Reference(ws, min_col=val_col, min_row=chart_data_row_start,
                         max_row=chart_data_row_start + num_slices - 1)
    cats_ref = Reference(ws, min_col=cat_col, min_row=chart_data_row_start,
                         max_row=chart_data_row_start + num_slices - 1)
    chart.add_data(data_ref)
    chart.set_categories(cats_ref)
    chart.dataLabels = DataLabelList(showPercent=True)
    ws.add_chart(chart, chart_anchor)


# ─────────────────────────────────────────
# XLSX — COURSE OFFERING REPORT
# ─────────────────────────────────────────
def generate_offering_filtered_xlsx(offering, rows, aggregates, filename, filter_meta):
    rows = sorted(rows, key=lambda r: r['full_name'].lower())
    filter_meta = filter_meta or {}
    period_label, period_type = format_date_range(filter_meta)
    student_label = filter_meta.get('student_label', 'All Students')

    wb = openpyxl.Workbook()

    # ═══ Sheet 1: Report ═══════════════════
    ws = wb.active
    ws.title = "Attendance Report"
    ws.sheet_view.showGridLines = False

    # Title block (row 1-2)
    ws.merge_cells('A1:J1')
    ws['A1'].value = "Ethiopian Aviation University — Student Attendance Management System"
    ws['A1'].font = _xl_font(bold=True, color=XL_HEADER_FONT, size=13)
    ws['A1'].fill = _xl_fill(XL_HEADER_BG)
    ws['A1'].alignment = _xl_center()

    ws.merge_cells('A2:J2')
    ws['A2'].value = f"{offering.course.name}  |  Section {offering.section.name} (Year {offering.section.year})"
    ws['A2'].font = _xl_font(bold=True, color='FFFFFF', size=11)
    ws['A2'].fill = _xl_fill(XL_SUBHEADER)
    ws['A2'].alignment = _xl_center()
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    # Info block
    try:
        programme_name = offering.section.programme.name
    except Exception:
        programme_name = "N/A"
    try:
        sem = offering.section.semester
        semester_label = str(sem)  # uses Semester.__str__: "2024/25 - Sem 2"
    except Exception:
        semester_label = "N/A"

    info = [
        ("Programme:",        programme_name),
        ("Semester:",         semester_label),
        ("Report Period:",    period_label),
        ("Filter Type:",      period_type),
        ("Student Filter:",   student_label),
        ("Credit Hours:",     str(offering.course.total_credit_hours)),
        ("Min Required Hrs:", str(offering.course.minimum_required_hours)),
        ("Generated:",        date.today().strftime('%d %B %Y')),
    ]
    current_row = 4
    for label, val in info:
        ws.cell(current_row, 1).value = label
        ws.cell(current_row, 1).font = _xl_font(bold=True, color=XL_SUBHEADER)
        ws.cell(current_row, 2).value = val
        ws.cell(current_row, 2).font = _xl_font(color='1F2937')
        ws.row_dimensions[current_row].height = 15
        current_row += 1

    current_row += 1

    # Count row
    total = aggregates['total_students']
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    ws.cell(current_row, 1).value = f"Total Students in this Report: {total}"
    ws.cell(current_row, 1).fill = _xl_fill(XL_HEADER_BG)
    ws.cell(current_row, 1).font = _xl_font(bold=True, color='FFFFFF', size=11)
    ws.cell(current_row, 1).alignment = _xl_center()
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    # Table header
    headers = ['#', 'Student ID', 'Full Name', 'Present Hrs', 'Late Hrs',
               'Excused Hrs', 'Absent Hrs', 'Attended %', 'Min Req. Hrs', 'Status',
               'Late Dates', 'Absent Dates']
    _xl_apply_table_header(ws, current_row, headers)
    ws.row_dimensions[current_row].height = 18
    table_header_row = current_row
    current_row += 1

    # Data rows
    for idx, row in enumerate(rows, start=1):
        alt = (idx % 2 == 0)
        status_fill, status_font = _status_fill_font(row['status'])
        row_fill = _xl_fill(XL_ALT_ROW) if alt else _xl_fill('FFFFFF')

        late_dates_str = ', '.join(_short_date(d) for d in row.get('late_dates', [])) or '—'
        absent_dates_str = ', '.join(_short_date(d) for d in row.get('absent_dates', [])) or '—'

        data = [
            idx,
            row['student_id'],
            row['full_name'],
            round(row.get('present_hours', row.get('attended_hours', 0)), 1),
            round(row.get('late_hours', 0), 1),
            round(row.get('excused_hours', 0), 1),
            round(row.get('absent_hours', row.get('missed_hours', 0)), 1),
            f"{row['percentage']}%",
            row['minimum_required'],
            row['status'],
            late_dates_str,
            absent_dates_str,
        ]
        alignments = [_xl_center()] * 12
        alignments[2] = _xl_left()  # full name left-aligned
        alignments[10] = _xl_left()
        alignments[11] = _xl_left()

        for j, (val, align) in enumerate(zip(data, alignments), start=1):
            c = ws.cell(current_row, j)
            c.value = val
            c.alignment = align
            c.border = _xl_border()
            if j == 10:  # Status column
                c.fill = status_fill
                c.font = status_font
            else:
                c.fill = row_fill
                c.font = _xl_font(color='1F2937', size=9)
        ws.row_dimensions[current_row].height = 15
        current_row += 1

    # Column widths
    col_widths_map = {'A': 5, 'B': 14, 'C': 28, 'D': 12, 'E': 10,
                      'F': 12, 'G': 11, 'H': 12, 'I': 13, 'J': 12,
                      'K': 18, 'L': 18}
    for col, w in col_widths_map.items():
        ws.column_dimensions[col].width = w

    current_row += 1

    # ── KPI summary block ─────────────────
    current_row = _xl_write_section_title(ws, current_row, "📊 Summary Statistics", col_span=10)
    kpi_headers = ['Metric', 'Value']
    _xl_apply_table_header(ws, current_row, kpi_headers)
    current_row += 1

    kpis = [
        ("Total Students", total),
        ("Average Attendance %", f"{aggregates['average_attendance_percentage']}%"),
        ("Safe", aggregates['safe_count']),
        ("Warning", aggregates['warning_count']),
        ("At Risk", aggregates['at_risk_count']),
    ]
    for label, val in kpis:
        ws.cell(current_row, 1).value = label
        ws.cell(current_row, 1).font = _xl_font(bold=True, color='374151')
        ws.cell(current_row, 1).border = _xl_border()
        ws.cell(current_row, 1).alignment = _xl_left()
        ws.cell(current_row, 2).value = val
        ws.cell(current_row, 2).font = _xl_font(color='1F2937')
        ws.cell(current_row, 2).border = _xl_border()
        ws.cell(current_row, 2).alignment = _xl_center()
        current_row += 1

    current_row += 1

    # ═══ Sheet 2: Charts Data ══════════════
    ws2 = wb.create_sheet("Charts")
    ws2.sheet_view.showGridLines = False

    ws2['A1'].value = "Attendance Bands"
    ws2['A1'].font = _xl_font(bold=True, color=XL_SUBHEADER, size=11)

    # Legend explanation
    ws2.merge_cells('A2:B2')
    ws2['A2'].value = "Bands: <75% = At Risk  |  75–84.9% = Warning  |  85–89.9% = Near-Safe  |  ≥90% = Safe"
    ws2['A2'].font = _xl_font(bold=False, color='374151', size=9)
    ws2['A2'].alignment = _xl_center()

    band_labels = ['<75%', '75-84.9%', '85-89.9%', '>=90%']
    band_display = ['<75%', '75–84.9%', '85–89.9%', '≥90%']
    band_values = [
        aggregates['attendance_bands'].get('<75%', 0),
        aggregates['attendance_bands'].get('75-84.9%', 0),
        aggregates['attendance_bands'].get('85-89.9%', 0),
        aggregates['attendance_bands'].get('>=90%', 0),
    ]
    _xl_apply_table_header(ws2, 3, ['Band', 'Students'])
    for i, (label, val) in enumerate(zip(band_display, band_values)):
        ws2.cell(4 + i, 1).value = label
        ws2.cell(4 + i, 2).value = val
        ws2.cell(4 + i, 1).font = _xl_font()
        ws2.cell(4 + i, 2).font = _xl_font()

    ws2.cell(11, 1).value = "Risk Distribution"
    ws2.cell(11, 1).font = _xl_font(bold=True, color=XL_SUBHEADER, size=11)

    risk_labels = ['Safe', 'Warning', 'At Risk']
    risk_values = [aggregates['safe_count'], aggregates['warning_count'], aggregates['at_risk_count']]
    _xl_apply_table_header(ws2, 12, ['Status', 'Count'])
    for i, (label, val) in enumerate(zip(risk_labels, risk_values)):
        ws2.cell(13 + i, 1).value = label
        ws2.cell(13 + i, 2).value = val
        ws2.cell(13 + i, 1).font = _xl_font()
        ws2.cell(13 + i, 2).font = _xl_font()

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 12

    # Add charts to Sheet 2
    _xl_add_bar_chart(ws2, chart_data_row_start=4, num_categories=4,
                       chart_anchor="D3", title="Attendance Bands",
                       series_col=2, cat_col=1)
    _xl_add_pie_chart(ws2, chart_data_row_start=13, num_slices=3,
                       chart_anchor="D25", title="Risk Distribution",
                       val_col=2, cat_col=1)

    # ── Write to buffer ───────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────
# XLSX — STUDENT INDIVIDUAL REPORT
# ─────────────────────────────────────────
def generate_student_xlsx(student, course_summaries, filter_meta=None):
    filter_meta = filter_meta or {}
    period_label, period_type = format_date_range(filter_meta)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Report"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'].value = "Ethiopian Aviation University — Student Attendance Report"
    ws['A1'].font = _xl_font(bold=True, color='FFFFFF', size=13)
    ws['A1'].fill = _xl_fill(XL_HEADER_BG)
    ws['A1'].alignment = _xl_center()
    ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:G2')
    ws['A2'].value = f"{student.full_name}  |  Student ID: {student.student_id}"
    ws['A2'].font = _xl_font(bold=True, color='FFFFFF', size=11)
    ws['A2'].fill = _xl_fill(XL_SUBHEADER)
    ws['A2'].alignment = _xl_center()
    ws.row_dimensions[2].height = 18

    current_row = 4
    info = [
        ("Report Period:", period_label),
        ("Filter Type:",   period_type),
        ("Generated:",     date.today().strftime('%d %B %Y')),
    ]
    for label, val in info:
        ws.cell(current_row, 1).value = label
        ws.cell(current_row, 1).font = _xl_font(bold=True, color=XL_SUBHEADER)
        ws.cell(current_row, 2).value = val
        ws.cell(current_row, 2).font = _xl_font(color='1F2937')
        current_row += 1

    current_row += 1
    total = len(course_summaries)

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    ws.cell(current_row, 1).value = f"Number of Courses in this Report: {total}"
    ws.cell(current_row, 1).fill = _xl_fill(XL_HEADER_BG)
    ws.cell(current_row, 1).font = _xl_font(bold=True, color='FFFFFF', size=11)
    ws.cell(current_row, 1).alignment = _xl_center()
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    headers = ['#', 'Course', 'Attended Hrs', 'Missed Hrs', 'Attendance %', 'Min Req. Hrs', 'Status']
    _xl_apply_table_header(ws, current_row, headers)
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    for idx, cs in enumerate(course_summaries, start=1):
        alt = (idx % 2 == 0)
        status_fill, status_font = _status_fill_font(cs['status'])
        row_fill = _xl_fill(XL_ALT_ROW) if alt else _xl_fill('FFFFFF')
        data = [
            idx, cs['course_name'],
            round(cs['attended_hours'], 1),
            round(cs['missed_hours'], 1),
            f"{cs['percentage']}%",
            cs['minimum_required'],
            cs['status'],
        ]
        aligns = [_xl_center(), _xl_left()] + [_xl_center()] * 5
        for j, (val, align) in enumerate(zip(data, aligns), start=1):
            c = ws.cell(current_row, j)
            c.value = val
            c.alignment = align
            c.border = _xl_border()
            if j == 7:
                c.fill = status_fill
                c.font = status_font
            else:
                c.fill = row_fill
                c.font = _xl_font(color='1F2937', size=9)
        ws.row_dimensions[current_row].height = 15
        current_row += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12

    # ── Late & Absent Days sheet ──────────────────────────────────────
    day_rows = []
    for cs in course_summaries:
        for d in cs.get('late_dates', []):
            day_rows.append((d, cs['course_name'], 'Late'))
        for d in cs.get('absent_dates', []):
            day_rows.append((d, cs['course_name'], 'Absent'))
    day_rows.sort(key=lambda r: r[0])

    if day_rows:
        ws3 = wb.create_sheet("Late & Absent Days")
        ws3.sheet_view.showGridLines = False
        ws3.merge_cells('A1:C1')
        ws3['A1'].value = "Days Marked Late or Absent"
        ws3['A1'].font = _xl_font(bold=True, color='FFFFFF', size=12)
        ws3['A1'].fill = _xl_fill(XL_HEADER_BG)
        ws3['A1'].alignment = _xl_center()
        ws3.row_dimensions[1].height = 20

        _xl_apply_table_header(ws3, 2, ['Date', 'Course', 'Status'])
        for idx, (d, course_name, day_status) in enumerate(day_rows, start=3):
            alt = (idx % 2 == 1)
            row_fill = _xl_fill(XL_ALT_ROW) if alt else _xl_fill('FFFFFF')
            status_fill, status_font = _status_fill_font(
                'At Risk' if day_status == 'Absent' else 'Warning')
            try:
                d_display = str(d)
            except Exception:
                d_display = str(d)
            cells = [d_display, course_name, day_status]
            aligns = [_xl_center(), _xl_left(), _xl_center()]
            for j, (val, align) in enumerate(zip(cells, aligns), start=1):
                c = ws3.cell(idx, j)
                c.value = val
                c.alignment = align
                c.border = _xl_border()
                if j == 3:
                    c.fill = status_fill
                    c.font = status_font
                else:
                    c.fill = row_fill
                    c.font = _xl_font(color='1F2937', size=9)
        ws3.column_dimensions['A'].width = 16
        ws3.column_dimensions['B'].width = 32
        ws3.column_dimensions['C'].width = 12

    # Charts sheet
    ws2 = wb.create_sheet("Risk Chart")
    ws2.sheet_view.showGridLines = False
    safe    = sum(1 for c in course_summaries if c['status'] == 'Safe')
    warning = sum(1 for c in course_summaries if c['status'] == 'Warning')
    at_risk = sum(1 for c in course_summaries if c['status'] == 'At Risk')

    ws2.cell(1, 1).value = "Status"
    ws2.cell(1, 2).value = "Count"
    ws2.cell(1, 1).font = _xl_font(bold=True, color=XL_HEADER_FONT)
    ws2.cell(1, 1).fill = _xl_fill(XL_HEADER_BG)
    ws2.cell(1, 2).font = _xl_font(bold=True, color=XL_HEADER_FONT)
    ws2.cell(1, 2).fill = _xl_fill(XL_HEADER_BG)
    for i, (label, val) in enumerate([('Safe', safe), ('Warning', warning), ('At Risk', at_risk)]):
        ws2.cell(2 + i, 1).value = label
        ws2.cell(2 + i, 2).value = val

    _xl_add_pie_chart(ws2, 2, 3, "D2", title="Course Risk Distribution")
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 10

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="student_{student.student_id}_report.xlsx"'
    )
    return response


# ─────────────────────────────────────────
# XLSX — SUMMARY OVERVIEW
# ─────────────────────────────────────────
def generate_summary_overview_xlsx(payload, filename):
    wb = openpyxl.Workbook()

    # ═══ Sheet 1: Overview ════════════════
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:G1')
    ws['A1'].value = "EAU — Attendance Executive Summary"
    ws['A1'].font = _xl_font(bold=True, color='FFFFFF', size=14)
    ws['A1'].fill = _xl_fill(XL_HEADER_BG)
    ws['A1'].alignment = _xl_center()
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A2:G2')
    ws['A2'].value = f"Generated: {date.today().strftime('%d %B %Y')}"
    ws['A2'].font = _xl_font(color='FFFFFF', size=10)
    ws['A2'].fill = _xl_fill(XL_SUBHEADER)
    ws['A2'].alignment = _xl_center()
    ws.row_dimensions[2].height = 16

    current_row = 4
    current_row = _xl_write_section_title(ws, current_row, "Key Performance Indicators", col_span=7)
    _xl_apply_table_header(ws, current_row, ['Metric', 'Value'])
    current_row += 1

    kpi = payload.get('kpis', {})
    for label, val in [
        ("Total Offerings",         kpi.get('total_offerings', 0)),
        ("Total Students",          kpi.get('total_students', 0)),
        ("Overall Average Att. %",  f"{kpi.get('overall_average_attendance', 0)}%"),
        ("Total At Risk Students",  kpi.get('total_at_risk_students', 0)),
        ("Worst Offering",          kpi.get('worst_offering_name', 'N/A')),
    ]:
        ws.cell(current_row, 1).value = label
        ws.cell(current_row, 1).font = _xl_font(bold=True, color='374151')
        ws.cell(current_row, 1).border = _xl_border()
        ws.cell(current_row, 1).alignment = _xl_left()
        ws.cell(current_row, 2).value = val
        ws.cell(current_row, 2).font = _xl_font(color='1F2937')
        ws.cell(current_row, 2).border = _xl_border()
        ws.cell(current_row, 2).alignment = _xl_center()
        current_row += 1

    current_row += 1
    current_row = _xl_write_section_title(ws, current_row, "Offering Analytics (At-Risk Hotspots)", col_span=7)

    count_row = current_row
    offerings = payload.get('offering_analytics', [])
    ws.merge_cells(start_row=count_row, start_column=1, end_row=count_row, end_column=7)
    ws.cell(count_row, 1).value = f"Total Offerings in this Report: {len(offerings)}"
    ws.cell(count_row, 1).fill = _xl_fill(XL_HEADER_BG)
    ws.cell(count_row, 1).font = _xl_font(bold=True, color='FFFFFF')
    ws.cell(count_row, 1).alignment = _xl_center()
    ws.row_dimensions[count_row].height = 16
    current_row += 1

    headers = ['Offering', 'Programme', 'Department', 'Students', 'Avg %', 'At Risk', 'Trend']
    _xl_apply_table_header(ws, current_row, headers)
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    for idx, row in enumerate(offerings):
        alt = (idx % 2 == 0)
        at_risk_fill = _xl_fill(XL_RISK_BG) if row.get('at_risk_count', 0) > 0 else _xl_fill('FFFFFF' if not alt else XL_ALT_ROW)
        data = [
            row.get('offering_label', ''),
            row.get('programme_name', ''),
            row.get('department_name', ''),
            row.get('student_count', 0),
            f"{row.get('average_attendance', 0)}%",
            row.get('at_risk_count', 0),
            row.get('trend_delta', 0),
        ]
        for j, val in enumerate(data, start=1):
            c = ws.cell(current_row, j)
            c.value = val
            c.border = _xl_border()
            c.alignment = _xl_left() if j <= 3 else _xl_center()
            c.font = _xl_font(color='1F2937', size=9)
            if j == 6 and row.get('at_risk_count', 0) > 0:
                c.fill = at_risk_fill
                c.font = _xl_font(bold=True, color=XL_RISK_FG, size=9)
            else:
                c.fill = _xl_fill(XL_ALT_ROW) if alt else _xl_fill('FFFFFF')
        ws.row_dimensions[current_row].height = 14
        current_row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10

    # ═══ Sheet 2: Attendance Bands ════════
    ws_bands = wb.create_sheet("Attendance Bands")
    ws_bands.sheet_view.showGridLines = False
    ws_bands.merge_cells('A1:B1')
    ws_bands['A1'].value = "Attendance Bands Distribution"
    ws_bands['A1'].font = _xl_font(bold=True, color=XL_HEADER_FONT, size=11)
    ws_bands['A1'].fill = _xl_fill(XL_HEADER_BG)
    ws_bands['A1'].alignment = _xl_center()

    # Legend explanation row
    ws_bands.merge_cells('A2:B2')
    ws_bands['A2'].value = "Bands: <75% = At Risk  |  75–84.9% = Warning  |  85–89.9% = Near-Safe  |  ≥90% = Safe"
    ws_bands['A2'].font = _xl_font(bold=False, color='374151', size=9)
    ws_bands['A2'].alignment = _xl_center()

    _xl_apply_table_header(ws_bands, 3, ['Band', 'No. of Students'])
    bands_start = 4
    band_labels_summary = ['<75%', '75-84.9%', '85-89.9%', '>=90%']
    # Aggregate bands across all offerings
    agg_bands: dict = {}
    for ofrow in offerings:
        for band, cnt in ofrow.get('attendance_bands', {}).items():
            agg_bands[band] = agg_bands.get(band, 0) + cnt
    # Also try top-level if provided
    if not agg_bands:
        top_bands = payload.get('attendance_bands', {})
        agg_bands = dict(top_bands)
    for i, lbl in enumerate(band_labels_summary):
        ws_bands.cell(bands_start + i, 1).value = lbl
        ws_bands.cell(bands_start + i, 2).value = agg_bands.get(lbl, 0)
        ws_bands.cell(bands_start + i, 1).font = _xl_font()
        ws_bands.cell(bands_start + i, 2).font = _xl_font()
    _xl_add_bar_chart(ws_bands, bands_start, 4, "D3", title="Attendance Bands", series_col=2, cat_col=1)
    ws_bands.column_dimensions['A'].width = 18
    ws_bands.column_dimensions['B'].width = 18

    # ═══ Sheet 3: Trend ═══════════════════
    ws_trend = wb.create_sheet("Attendance Trend")
    ws_trend.sheet_view.showGridLines = False
    ws_trend.merge_cells('A1:B1')
    ws_trend['A1'].value = "Attendance Trend by Week"
    ws_trend['A1'].font = _xl_font(bold=True, color=XL_HEADER_FONT, size=11)
    ws_trend['A1'].fill = _xl_fill(XL_HEADER_BG)
    ws_trend['A1'].alignment = _xl_center()

    _xl_apply_table_header(ws_trend, 2, ['Period', 'Average Attendance %'])
    trend_start = 3
    trend_data = payload.get('attendance_trend', [])
    for i, row in enumerate(trend_data):
        ws_trend.cell(trend_start + i, 1).value = row.get('period', '')
        ws_trend.cell(trend_start + i, 2).value = row.get('average_attendance', 0)
        ws_trend.cell(trend_start + i, 1).font = _xl_font()
        ws_trend.cell(trend_start + i, 2).font = _xl_font()

    if trend_data:
        _xl_add_bar_chart(ws_trend, trend_start, len(trend_data), "D2",
                           title="Weekly Attendance Trend", series_col=2, cat_col=1)
    ws_trend.column_dimensions['A'].width = 16
    ws_trend.column_dimensions['B'].width = 22

    # ═══ Sheet 3: Risk Charts ══════════════
    ws_risk = wb.create_sheet("Risk Distribution")
    ws_risk.sheet_view.showGridLines = False
    _xl_apply_table_header(ws_risk, 1, ['Status', 'Count'])
    risk_dist = payload.get('risk_distribution', {})
    for i, (label, val) in enumerate(risk_dist.items()):
        ws_risk.cell(2 + i, 1).value = label
        ws_risk.cell(2 + i, 2).value = val
    if risk_dist:
        _xl_add_pie_chart(ws_risk, 2, len(risk_dist), "D2", title="Risk Distribution")
    ws_risk.column_dimensions['A'].width = 14
    ws_risk.column_dimensions['B'].width = 10

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fn = filename.replace('.csv', '.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{fn}"'
    return response


# ─────────────────────────────────────────
# LEGACY CSV WRAPPERS (kept for compat)
# ─────────────────────────────────────────
def generate_course_csv(course, summary, filename):
    """Legacy — now delegates to XLSX."""
    # Build a minimal filter_meta and aggregates for the xlsx generator
    summary = sorted(summary, key=lambda r: r['full_name'].lower())
    safe    = sum(1 for r in summary if r['status'] == 'Safe')
    warning = sum(1 for r in summary if r['status'] == 'Warning')
    at_risk = sum(1 for r in summary if r['status'] == 'At Risk')
    bands = {'<75%': 0, '75-84.9%': 0, '85-89.9%': 0, '>=90%': 0}
    for r in summary:
        p = r['percentage']
        if p < 75:    bands['<75%'] += 1
        elif p < 85:  bands['75-84.9%'] += 1
        elif p < 90:  bands['85-89.9%'] += 1
        else:         bands['>=90%'] += 1
    aggregates = {
        'total_students': len(summary), 'average_attendance_percentage': 0,
        'safe_count': safe, 'warning_count': warning, 'at_risk_count': at_risk,
        'risk_distribution': {'Safe': safe, 'Warning': warning, 'At Risk': at_risk},
        'attendance_bands': bands,
    }
    # Build a fake offering with just the course
    class FakeOffering:
        pass
    fo = FakeOffering()
    fo.course = course
    fo.section = type('S', (), {'name': 'N/A', 'year': 'N/A',
                                 'programme': type('P', (), {'name': 'N/A'})(),
                                 'semester': type('SM', (), {'label': 'N/A'})()})()
    return generate_offering_filtered_xlsx(
        fo, summary, aggregates,
        filename.replace('.csv', '.xlsx'), {}
    )


def generate_student_csv(student, course_summaries, filter_meta=None):
    """Legacy — delegates to XLSX."""
    return generate_student_xlsx(student, course_summaries, filter_meta)


def generate_offering_filtered_csv(offering, rows, aggregates, filename, filter_meta):
    """Main CSV entry point — delegates to XLSX."""
    return generate_offering_filtered_xlsx(offering, rows, aggregates,
                                            filename.replace('.csv', '.xlsx'), filter_meta)


def generate_summary_overview_csv(payload, filename):
    """Main summary CSV entry point — delegates to XLSX."""
    return generate_summary_overview_xlsx(payload, filename.replace('.csv', '.xlsx'))


# ─────────────────────────────────────────
# SUMMARY PDF
# ─────────────────────────────────────────
def generate_summary_overview_pdf(payload, title="Executive Attendance Summary"):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    elements = []
    styles = getSampleStyleSheet()
    build_pdf_header(elements, title, "Cross-offering analytics snapshot", styles)

    kpi = payload.get('kpis', {})
    kpi_data = [['Metric', 'Value']]
    for label, val in [
        ("Total Offerings",         str(kpi.get('total_offerings', 0))),
        ("Total Students",          str(kpi.get('total_students', 0))),
        ("Overall Average Att. %",  f"{kpi.get('overall_average_attendance', 0)}%"),
        ("Total At Risk Students",  str(kpi.get('total_at_risk_students', 0))),
        ("Worst Offering",          kpi.get('worst_offering_name', 'N/A')),
    ]:
        kpi_data.append([label, val])

    kpi_table = Table(kpi_data, colWidths=[220, 420])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), DARK_GREEN),
        ('TEXTCOLOR', (0,0), (-1,0), WHITE),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [WHITE, LIGHT_GREEN]),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 14))

    # ── Charts row: Attendance Bands + Risk Distribution ────────────────
    risk_dist = payload.get('risk_distribution', {})
    agg_bands = payload.get('attendance_bands', {})

    # Build attendance bands aggregated from offering_analytics if not provided directly
    if not agg_bands:
        for ofrow in payload.get('offering_analytics', []):
            for band, cnt in ofrow.get('attendance_bands', {}).items():
                agg_bands[band] = agg_bands.get(band, 0) + cnt

    band_chart_data = []
    for lbl in ['<75%', '75-84.9%', '85-89.9%', '>=90%']:
        display = lbl.replace('>=', '\u226590%').replace('75-', '75\u201384.9%').replace('85-', '85\u201389.9%') if lbl != '<75%' else '<75%'
        display = lbl
        band_chart_data.append((display, agg_bands.get(lbl, 0)))

    bar_d = build_pdf_bar_chart(band_chart_data, width=310, height=185) if any(v for _, v in band_chart_data) else None
    pie_d = None
    if risk_dist:
        risk_data = list(risk_dist.items())
        pie_d = build_pdf_pie_chart(risk_data, width=260, height=185)

    if bar_d and pie_d:
        chart_row = Table([[bar_d, pie_d]], colWidths=[350, 290])
        chart_row.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
        elements.append(chart_row)
        elements.append(Spacer(1, 10))
    elif pie_d:
        elements.append(pie_d)
        elements.append(Spacer(1, 8))
    elif bar_d:
        elements.append(bar_d)
        elements.append(Spacer(1, 8))

    # ── Attendance Trend bar chart ──────────────────────────────────────
    trend_data = payload.get('attendance_trend', [])
    if trend_data:
        elements.append(Paragraph("Attendance Trend by Week", ParagraphStyle(
            'h3', fontSize=11, fontName='Helvetica-Bold',
            textColor=DARK_GREEN, spaceAfter=5)))
        trend_band = [(row.get('period', ''), row.get('average_attendance', 0)) for row in trend_data]
        trend_d = build_pdf_bar_chart(trend_band, width=620, height=185)
        if trend_d:
            elements.append(trend_d)
            elements.append(Spacer(1, 10))

    # Offering analytics table
    offerings = payload.get('offering_analytics', [])[:15]
    if offerings:
        total_offerings = len(offerings)
        count_data = [[f"Total Offerings in this Report: {total_offerings}"]]
        count_table = Table(count_data, colWidths=[640])
        count_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), DARK_GREEN),
            ('TEXTCOLOR', (0,0), (-1,-1), WHITE),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(count_table)

        offering_data = [['Offering', 'Programme', 'Dept', 'Students', 'Avg %', 'At Risk', 'Trend']]
        for row in offerings:
            offering_data.append([
                row.get('offering_label', ''),
                row.get('programme_name', ''),
                row.get('department_name', ''),
                row.get('student_count', 0),
                f"{row.get('average_attendance', 0)}%",
                row.get('at_risk_count', 0),
                row.get('trend_delta', 0),
            ])
        offering_table = Table(offering_data, colWidths=[210, 130, 100, 60, 55, 55, 55], repeatRows=1)
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), DARK_GREEN),
            ('TEXTCOLOR', (0,0), (-1,0), WHITE),
            ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [WHITE, LIGHT_GREEN]),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (0,1), (0,-1), 'LEFT'),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ])
        for i, row in enumerate(offerings, start=1):
            if row.get('at_risk_count', 0) > 0:
                style.add('TEXTCOLOR', (5, i), (5, i), RED)
                style.add('FONTNAME',  (5, i), (5, i), 'Helvetica-Bold')
        offering_table.setStyle(style)
        elements.append(Spacer(1, 8))
        elements.append(Paragraph("At-Risk Hotspots", ParagraphStyle(
            'h3', fontSize=11, fontName='Helvetica-Bold',
            textColor=DARK_GREEN, spaceAfter=5)))
        elements.append(offering_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ─────────────────────────────────────────
# DATA HELPERS (used by views.py directly)
# ─────────────────────────────────────────
def get_offering_student_report_data(
    offering,
    start_date=None,
    end_date=None,
    student_id=None,
    warning_threshold=85.0,
    at_risk_threshold=75.0,
):
    enrollments = Enrollment.objects.filter(
        section=offering.section,
        status='active',
    ).select_related('student')

    if student_id:
        enrollments = enrollments.filter(student_id=student_id)

    rows = []
    for enrollment in enrollments:
        student = enrollment.student
        records = AttendanceRecord.objects.filter(
            student=student,
            course_offering=offering,
        )
        if start_date:
            records = records.filter(date__gte=start_date)
        if end_date:
            records = records.filter(date__lte=end_date)

        present_hours = records.filter(status='present').aggregate(
            total=Sum('hours_attended'))['total'] or Decimal('0')
        late_hours = records.filter(status='late').aggregate(
            total=Sum('hours_attended'))['total'] or Decimal('0')
        excused_qs = records.filter(status='excused')
        absent_qs  = records.filter(status='absent')
        excused_hours = excused_qs.aggregate(total=Sum('hours_attended'))['total'] or Decimal('0')
        absent_hours  = absent_qs.aggregate(total=Sum('hours_attended'))['total']  or Decimal('0')

        if excused_qs.exists() and excused_hours == 0:
            fallback = records.filter(status__in=['present','late'],
                                      hours_attended__gt=0).aggregate(
                avg=Avg('hours_attended'))['avg'] or Decimal('1.0')
            excused_hours = fallback * Decimal(excused_qs.count())
        if absent_qs.exists() and absent_hours == 0:
            fallback = records.filter(status__in=['present','late'],
                                      hours_attended__gt=0).aggregate(
                avg=Avg('hours_attended'))['avg'] or Decimal('1.0')
            absent_hours = fallback * Decimal(absent_qs.count())

        # Excused excluded from denominator — not penalised, not credited
        attended_hours   = present_hours + late_hours
        missed_hours     = absent_hours   # excused NOT counted as missed
        total_logged     = attended_hours + missed_hours
        total_credit     = _D(str(offering.course.total_credit_hours))
        effective_credit = max(total_credit - excused_hours, _D('1'))
        late_count       = records.filter(status='late').count()
        earned, percentage = calc_attendance(present_hours, late_hours, effective_credit, late_count)

        if not is_gate_open(total_logged, total_credit):
            status = 'Safe'
        elif percentage < at_risk_threshold:
            status = 'At Risk'
        elif percentage < warning_threshold:
            status = 'Warning'
        else:
            status = 'Safe'

        # Display: each late session = 0.5hr "late", rest credited as "present"
        total_late_penalty = LATE_DEDUCTION * _D(str(late_count))
        late_penalty    = min(late_hours, total_late_penalty)
        late_earned     = late_hours - late_penalty
        display_present = present_hours + late_earned
        display_late    = late_penalty

        # Collect excused comments for report display (safe if migration not yet run)
        try:
            excused_comments = list(excused_qs.exclude(comment='').values_list('comment', flat=True))
            excused_reason   = '; '.join(excused_comments) if excused_comments else ''
        except Exception:
            excused_reason = ''

        # Day-by-day dates of late/absent sessions (for "which day were they
        # late/absent" columns in reports)
        late_dates = list(
            records.filter(status='late').order_by('date')
                   .values_list('date', flat=True)
        )
        absent_dates = list(
            records.filter(status='absent').order_by('date')
                   .values_list('date', flat=True)
        )

        rows.append({
            'student_pk':       student.id,
            'student_id':       student.student_id,
            'full_name':        student.full_name,
            'present_hours':    float(display_present),
            'late_hours':       float(display_late),
            'excused_hours':    float(excused_hours),
            'absent_hours':     float(absent_hours),
            'attended_hours':   float(earned),
            'missed_hours':     float(missed_hours),
            'total_hours':      float(total_credit),
            'percentage':       percentage,
            'minimum_required': float(offering.course.minimum_required_hours),
            'status':           status,
            'excused_reason':   excused_reason,
            'late_dates':       [d.isoformat() for d in late_dates],
            'absent_dates':     [d.isoformat() for d in absent_dates],
        })

    return rows


def build_offering_report_aggregates(rows):
    total_students = len(rows)
    safe_count    = sum(1 for r in rows if r['status'] == 'Safe')
    warning_count = sum(1 for r in rows if r['status'] == 'Warning')
    at_risk_count = sum(1 for r in rows if r['status'] == 'At Risk')

    avg_percentage = round(
        sum(r['percentage'] for r in rows) / total_students, 1
    ) if total_students else 0.0

    bands = {'<75%': 0, '75-84.9%': 0, '85-89.9%': 0, '>=90%': 0}
    for row in rows:
        pct = row['percentage']
        if pct < 75:    bands['<75%'] += 1
        elif pct < 85:  bands['75-84.9%'] += 1
        elif pct < 90:  bands['85-89.9%'] += 1
        else:           bands['>=90%'] += 1

    return {
        'total_students':                total_students,
        'average_attendance_percentage': avg_percentage,
        'safe_count':                    safe_count,
        'warning_count':                 warning_count,
        'at_risk_count':                 at_risk_count,
        'risk_distribution': {
            'Safe': safe_count, 'Warning': warning_count, 'At Risk': at_risk_count,
        },
        'attendance_bands': bands,
    }