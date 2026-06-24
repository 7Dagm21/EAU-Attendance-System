"""
attendance_excel.py — Two views for Excel-based attendance workflow:

1. AttendanceTemplateView  GET  /api/attendance/template/<offering_id>/
   Downloads an .xlsx template pre-filled with enrolled students
   and date columns for a selected week.

2. AttendanceImportView    POST /api/attendance/import/
   Parses the uploaded .xlsx and submits one AttendanceRecord
   per student per date column that has data.
"""

import io
from datetime import date, timedelta
from decimal import Decimal

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import (
    CourseOffering, Enrollment, Student,
    AttendanceRecord, SystemSettings
)
from .utils import send_absence_alert, send_threshold_warning

# ── colours ───────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1A4A0F"   # dark green  — locked header cells
C_HEADER_FG  = "FFFFFF"   # white text
C_DATE_BG    = "E8F5E0"   # light green — date header cells
C_DATE_FG    = "1A4A0F"
C_INPUT_BG   = "FFFDE7"   # pale yellow — teacher fills these
C_LOCKED_BG  = "F5F5F5"   # light grey  — read-only (name / id cols)
C_LOCKED_FG  = "374151"
C_SUM_BG     = "DBEAFE"   # light blue  — summary cols

# ── helpers ───────────────────────────────────────────────────────────────────
def _thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(name='Arial', bold=bold, color=color, size=size)

def _week_dates(week_start: date):
    """Return Mon–Fri dates for the week containing week_start."""
    # Snap to Monday
    monday = week_start - timedelta(days=week_start.weekday())
    return [monday + timedelta(days=i) for i in range(5)]

STATUS_MAP = {
    # full words
    'present': 'present', 'late': 'late',
    'excused': 'excused', 'absent': 'absent',
    # single letters
    'p': 'present', 'l': 'late', 'e': 'excused', 'a': 'absent',
}

# ── Template Download ─────────────────────────────────────────────────────────
class AttendanceTemplateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, offering_id):
        try:
            offering = CourseOffering.objects.select_related(
                'course', 'section__programme', 'section__semester'
            ).get(id=offering_id)
        except CourseOffering.DoesNotExist:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

        # Date range — teacher can pass ?week_start=YYYY-MM-DD or we use current week
        week_start_str = request.query_params.get('week_start')
        custom_start   = request.query_params.get('start_date')
        custom_end     = request.query_params.get('end_date')

        if custom_start and custom_end:
            try:
                start = date.fromisoformat(custom_start)
                end   = date.fromisoformat(custom_end)
                all_days = []
                d = start
                while d <= end and len(all_days) < 31:  # allow up to 31 days for monthly
                    if d.weekday() < 5:
                        all_days.append(d)
                    d += timedelta(days=1)
            except ValueError:
                return Response({'error': 'Invalid date format. Use YYYY-MM-DD'},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            period = request.query_params.get('period', 'week')  # 'week' or 'month'
            if week_start_str:
                try:
                    week_start = date.fromisoformat(week_start_str)
                except ValueError:
                    return Response({'error': 'Invalid week_start. Use YYYY-MM-DD'},
                                    status=status.HTTP_400_BAD_REQUEST)
            else:
                week_start = date.today()
            if period == 'month':
                # 4 weeks starting from the Monday of the given week
                monday = week_start - timedelta(days=week_start.weekday())
                all_days = []
                for week_offset in range(4):
                    for day_offset in range(5):
                        all_days.append(monday + timedelta(weeks=week_offset, days=day_offset))
            else:
                all_days = _week_dates(week_start)
            all_days = _week_dates(week_start)

        # Get enrolled students
        enrollments = Enrollment.objects.filter(
            section=offering.section, status='active'
        ).select_related('student').order_by(
            'student__last_name', 'student__first_name'
        )
        students = [e.student for e in enrollments]

        if not students:
            return Response({'error': 'No enrolled students found'},
                            status=status.HTTP_400_BAD_REQUEST)

        # ── Build workbook ────────────────────────────────────────────────────
        wb = openpyxl.Workbook()

        # ── Instructions sheet ────────────────────────────────────────────────
        ins = wb.active
        ins.title = 'Instructions'
        ins.column_dimensions['A'].width = 80

        instructions = [
            ("EAU Student Attendance Management System", True, 14, C_HEADER_BG),
            ("Attendance Import Template — Instructions", False, 11, "374151"),
            ("", False, 10, "000000"),
            ("HOW TO FILL THIS TEMPLATE:", True, 11, "1A4A0F"),
            ("1. Go to the 'Attendance' sheet (tab at the bottom).", False, 10, "000000"),
            ("2. For each student and each date column, enter one of:", False, 10, "000000"),
            ("   P  or  present   = Student was present", False, 10, "000000"),
            ("   L  or  late      = Student arrived late", False, 10, "000000"),
            ("   E  or  excused   = Absence is excused", False, 10, "000000"),
            ("   A  or  absent    = Student was absent (unexcused)", False, 10, "000000"),
            ("   (Leave blank to skip that date — it won't be imported)", False, 10, "666666"),
            ("", False, 10, "000000"),
            ("3. You can also use dropdowns — click any yellow cell to see the options.", False, 10, "000000"),
            ("4. The 'Hours' column calculates AUTOMATICALLY based on P/L entries.", False, 10, "1A4A0F"),
            ("   P (present) = 1 hr.  L (late) = 0.5 hr.  Do NOT edit Hours manually.", False, 10, "000000"),
            ("5. DO NOT change student names, IDs, or column headers.", False, 10, "CC0000"),
            ("6. DO NOT add or remove rows or columns.", False, 10, "CC0000"),
            ("7. Save the file and upload it in the Teacher Portal.", False, 10, "000000"),
            ("", False, 10, "000000"),
            ("VALID STATUS VALUES:", True, 11, "1A4A0F"),
            ("Full word:    present  /  late  /  excused  /  absent", False, 10, "000000"),
            ("Short letter: P  /  L  /  E  /  A  (case-insensitive)", False, 10, "000000"),
        ]

        for row_data in instructions:
            text, bold, size, color = row_data
            cell = ins.cell(row=ins.max_row + 1 if ins.max_row > 0 else 1,
                            column=1, value=text)
            cell.font = Font(name='Arial', bold=bold, size=size, color=color)
            cell.alignment = Alignment(wrap_text=True)

        # ── Attendance sheet ──────────────────────────────────────────────────
        ws = wb.create_sheet('Attendance')

        # Meta info rows
        meta = [
            ("Course:",         offering.course.name),
            ("Section:",        f"{offering.section.name} — Year {offering.section.year}"),
            ("Programme:",      offering.section.programme.name),
            ("Semester:",       str(offering.section.semester)),
            ("Teacher:",        offering.teacher.get_full_name() if offering.teacher else "—"),
            ("Total Students:", str(len(students))),
            ("Generated:",      str(date.today())),
        ]
        for i, (label, value) in enumerate(meta, start=1):
            ws.cell(row=i, column=1, value=label).font = _font(bold=True, color=C_HEADER_BG)
            ws.cell(row=i, column=2, value=value).font = _font(color="374151")

        header_row = len(meta) + 2  # leave one blank row

        # Fixed columns: #, Student Name, Student ID
        FIXED_COLS = 3  # cols 1-3
        # Date columns start at col 4
        DATE_START_COL = FIXED_COLS + 1
        # After dates: Session Type, Hours
        n_dates      = len(all_days)
        SESSION_COL  = DATE_START_COL + n_dates
        HOURS_COL    = SESSION_COL + 1
        TOTAL_COLS   = HOURS_COL

        # ── Header row ────────────────────────────────────────────────────────
        headers_fixed = ["#", "Student Name", "Student ID"]
        for col, h in enumerate(headers_fixed, start=1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font      = _font(bold=True, color=C_HEADER_FG, size=10)
            cell.fill      = _fill(C_HEADER_BG)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = _thin_border()

        for i, d in enumerate(all_days):
            col  = DATE_START_COL + i
            cell = ws.cell(row=header_row, column=col,
                           value=d.strftime("%a\n%d/%m"))
            cell.font      = _font(bold=True, color=C_DATE_FG, size=9)
            cell.fill      = _fill(C_DATE_BG)
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
            cell.border    = _thin_border()
            # Store ISO date in a comment so the parser can read it back
            from openpyxl.comments import Comment
            cell.comment = Comment(d.isoformat(), "SAMS")

        for col, h in [(SESSION_COL, "Session Type"), (HOURS_COL, "Hours")]:
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font      = _font(bold=True, color=C_HEADER_FG, size=10)
            cell.fill      = _fill("2563EB")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = _thin_border()

        # Row height for header
        ws.row_dimensions[header_row].height = 32

        # ── Student rows ──────────────────────────────────────────────────────
        # Data validation — dropdown for status cells
        dv = DataValidation(
            type="list",
            formula1='"present,late,excused,absent,P,L,E,A"',
            allow_blank=True,
            showErrorMessage=True,
            error='Enter: present, late, excused, absent (or P, L, E, A)',
            errorTitle='Invalid status',
        )
        ws.add_data_validation(dv)

        dv_session = DataValidation(
            type="list",
            formula1='"theory,practical"',
            allow_blank=False,
            showErrorMessage=True,
        )
        ws.add_data_validation(dv_session)

        for idx, student in enumerate(students):
            row = header_row + 1 + idx

            # # col
            c = ws.cell(row=row, column=1, value=idx + 1)
            c.font = _font(color="9CA3AF"); c.fill = _fill(C_LOCKED_BG)
            c.alignment = Alignment(horizontal='center'); c.border = _thin_border()

            # Name col
            c = ws.cell(row=row, column=2, value=student.full_name)
            c.font = _font(bold=True, color=C_LOCKED_FG); c.fill = _fill(C_LOCKED_BG)
            c.alignment = Alignment(vertical='center'); c.border = _thin_border()

            # ID col
            c = ws.cell(row=row, column=3, value=student.student_id)
            c.font = _font(color=C_LOCKED_FG); c.fill = _fill(C_LOCKED_BG)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = _thin_border()

            # Date cells — yellow, teacher fills
            for i in range(n_dates):
                col  = DATE_START_COL + i
                cell = ws.cell(row=row, column=col, value="")
                cell.fill      = _fill(C_INPUT_BG)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border    = _thin_border()
                cell.font      = _font(size=10)
                dv.add(cell)

            # Session type — default theory
            c = ws.cell(row=row, column=SESSION_COL, value="theory")
            c.fill = _fill(C_SUM_BG); c.font = _font(size=10)
            c.alignment = Alignment(horizontal='center'); c.border = _thin_border()
            dv_session.add(c)

            # Hours — Excel formula: counts P/present/L/late cells × 1 hr each
            # Late counts as 0.5 hrs (deduction). Formula auto-updates as teacher fills.
            date_start_letter = get_column_letter(DATE_START_COL)
            date_end_letter   = get_column_letter(DATE_START_COL + n_dates - 1)
            hours_formula = (
                f'=COUNTIF({date_start_letter}{row}:{date_end_letter}{row},"P")'
                f'+COUNTIF({date_start_letter}{row}:{date_end_letter}{row},"present")'
                f'+COUNTIF({date_start_letter}{row}:{date_end_letter}{row},"L")*0.5'
                f'+COUNTIF({date_start_letter}{row}:{date_end_letter}{row},"late")*0.5'
            )
            c = ws.cell(row=row, column=HOURS_COL, value=hours_formula)
            c.fill = _fill(C_SUM_BG); c.font = _font(size=10, color="1A4A0F", bold=True)
            c.alignment = Alignment(horizontal='center'); c.border = _thin_border()

            ws.row_dimensions[row].height = 20

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions[get_column_letter(1)].width = 5   # #
        ws.column_dimensions[get_column_letter(2)].width = 28  # Name
        ws.column_dimensions[get_column_letter(3)].width = 14  # ID
        for i in range(n_dates):
            ws.column_dimensions[get_column_letter(DATE_START_COL + i)].width = 10
        ws.column_dimensions[get_column_letter(SESSION_COL)].width = 14
        ws.column_dimensions[get_column_letter(HOURS_COL)].width = 8

        # ── Freeze panes ──────────────────────────────────────────────────────
        ws.freeze_panes = ws.cell(row=header_row + 1,
                                   column=DATE_START_COL)

        # ── Metadata sheet (hidden, read by parser) ───────────────────────────
        meta_sheet = wb.create_sheet('_meta')
        meta_sheet['A1'] = 'offering_id'
        meta_sheet['B1'] = str(offering_id)
        meta_sheet['A2'] = 'header_row'
        meta_sheet['B2'] = str(header_row)
        meta_sheet['A3'] = 'date_start_col'
        meta_sheet['B3'] = str(DATE_START_COL)
        meta_sheet['A4'] = 'n_dates'
        meta_sheet['B4'] = str(n_dates)
        meta_sheet['A5'] = 'session_col'
        meta_sheet['B5'] = str(SESSION_COL)
        meta_sheet['A6'] = 'hours_col'
        meta_sheet['B6'] = str(HOURS_COL)
        # Store dates row
        for i, d in enumerate(all_days):
            meta_sheet.cell(row=7, column=i + 1, value=d.isoformat())
        meta_sheet.sheet_state = 'hidden'

        # ── Stream response ───────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name  = offering.course.name.replace(' ', '_')[:30]
        week_label = all_days[0].strftime('%Y%m%d') if all_days else 'template'
        filename   = f"Attendance_{safe_name}_{week_label}.xlsx"

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ── Excel Import ──────────────────────────────────────────────────────────────
class AttendanceImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes     = [MultiPartParser]

    def post(self, request):
        file        = request.FILES.get('file')
        preview_only = request.data.get('preview_only', 'false').lower() == 'true'

        if not file:
            return Response({'error': 'No file uploaded'},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
        except Exception as e:
            return Response({'error': f'Could not read Excel file: {e}'},
                            status=status.HTTP_400_BAD_REQUEST)

        # Read metadata
        try:
            meta  = wb['_meta']
            offering_id   = int(meta['B1'].value)
            header_row    = int(meta['B2'].value)
            date_start_col = int(meta['B3'].value)
            n_dates       = int(meta['B4'].value)
            session_col   = int(meta['B5'].value)
            hours_col     = int(meta['B6'].value)
            dates = [
                date.fromisoformat(str(meta.cell(row=7, column=i + 1).value))
                for i in range(n_dates)
            ]
        except Exception:
            return Response(
                {'error': 'This file was not generated by SAMS or has been corrupted. Please download a fresh template.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            offering = CourseOffering.objects.select_related(
                'course__programme', 'section'
            ).get(id=offering_id)
        except CourseOffering.DoesNotExist:
            return Response({'error': 'Course offering not found'},
                            status=status.HTTP_404_NOT_FOUND)

        ws = wb['Attendance']

        # Parse rows
        records   = []  # valid records to submit
        errors    = []  # parse problems
        skipped   = []  # blank cells

        data_start = header_row + 1
        for row_idx in range(data_start, ws.max_row + 1):
            student_id_val = ws.cell(row=row_idx, column=3).value
            if not student_id_val:
                break   # end of data

            student_id_str = str(student_id_val).strip()
            try:
                student = Student.objects.get(student_id=student_id_str)
            except Student.DoesNotExist:
                errors.append({
                    'row': row_idx,
                    'student_id': student_id_str,
                    'error': 'Student not found in system'
                })
                continue

            session_type_val = ws.cell(row=row_idx, column=session_col).value
            session_type     = str(session_type_val or 'theory').lower().strip()
            if session_type not in ('theory', 'practical'):
                session_type = 'theory'

            hours_val = ws.cell(row=row_idx, column=hours_col).value
            try:
                hours = float(hours_val or 1.5)
                hours = max(0.5, min(8.0, hours))
            except (ValueError, TypeError):
                hours = 1.5

            for i, att_date in enumerate(dates):
                col        = date_start_col + i
                status_val = ws.cell(row=row_idx, column=col).value

                if status_val is None or str(status_val).strip() == '':
                    skipped.append({'student_id': student_id_str, 'date': att_date.isoformat()})
                    continue

                normalized = STATUS_MAP.get(str(status_val).strip().lower())
                if not normalized:
                    errors.append({
                        'row': row_idx,
                        'student_id': student_id_str,
                        'date': att_date.isoformat(),
                        'value': str(status_val),
                        'error': f'Invalid status "{status_val}". Use: present/late/excused/absent or P/L/E/A'
                    })
                    continue

                records.append({
                    'student':      student,
                    'student_id':   student_id_str,
                    'student_name': student.full_name,
                    'date':         att_date,
                    'status':       normalized,
                    'session_type': session_type,
                    'hours':        Decimal(str(hours)) if normalized in ('present', 'late') else Decimal('0'),
                })

        # If preview only — return parsed data without saving
        if preview_only:
            return Response({
                'offering_id':    offering_id,
                'offering_name':  offering.course.name,
                'section':        offering.section.name,
                'dates':          [d.isoformat() for d in dates],
                'valid_count':    len(records),
                'skipped_count':  len(skipped),
                'error_count':    len(errors),
                'errors':         errors,
                'preview': [
                    {
                        'student_name': r['student_name'],
                        'student_id':   r['student_id'],
                        'date':         r['date'].isoformat(),
                        'status':       r['status'],
                        'session_type': r['session_type'],
                        'hours':        float(r['hours']),
                    }
                    for r in records[:100]  # cap preview at 100 rows
                ],
            })

        # Submit records
        teacher       = request.user
        created_count = 0
        updated_count = 0

        for r in records:
            obj, created = AttendanceRecord.objects.update_or_create(
                student=r['student'],
                course_offering=offering,
                date=r['date'],
                session_type=r['session_type'],
                defaults={
                    'status':         r['status'],
                    'hours_attended': r['hours'],
                    'recorded_by':    teacher,
                }
            )
            if created:
                created_count += 1
            else:
                updated_count += 1

            # Fire absence alerts
            if r['status'] == 'absent':
                try:
                    from .views import notify_elevated
                    teacher_name = teacher.get_full_name() or teacher.username
                    notify_elevated(
                        'absence',
                        f"{r['student_name']} (ID: {r['student_id']}) was absent in "
                        f"{offering.course.name} on {r['date']} (imported). By {teacher_name}.",
                        programme=offering.course.programme,
                    )
                except Exception:
                    pass

        return Response({
            'message':       f'{created_count} records created, {updated_count} updated.',
            'created':       created_count,
            'updated':       updated_count,
            'skipped':       len(skipped),
            'errors':        errors,
            'error_count':   len(errors),
        }, status=status.HTTP_201_CREATED)

class TeacherMonitoringView(APIView):
    """GET /api/teacher-monitoring/
    Returns for each teacher/offering: which scheduled class days (per
    TeachingSchedule) had no AttendanceRecord submitted. Allows admins
    to see if a teacher missed taking attendance."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from .models import TeachingSchedule, AttendanceRecord, CourseOffering
        from datetime import date, timedelta

        # Scope: admin sees all; teacher sees only their own
        if request.user.role == 'admin':
            offerings = CourseOffering.objects.select_related(
                'course', 'section', 'teacher'
            ).all()
        else:
            offerings = CourseOffering.objects.select_related(
                'course', 'section', 'teacher'
            ).filter(teacher=request.user)

        today = date.today()
        # Look back up to 30 days
        lookback_start = today - timedelta(days=30)

        result = []
        for offering in offerings:
            if not offering.teacher:
                continue
            slots = TeachingSchedule.objects.filter(course_offering=offering)
            if not slots.exists():
                continue

            # Build list of class days that should have attendance in the window
            expected_days = []
            d = lookback_start
            while d <= today:
                for slot in slots:
                    if d.weekday() == slot.day_of_week:
                        expected_days.append(d)
                        break
                d += timedelta(days=1)

            # Which of those days actually have records?
            submitted_dates = set(
                AttendanceRecord.objects.filter(
                    course_offering=offering,
                    date__in=expected_days,
                ).values_list('date', flat=True).distinct()
            )

            missed_days = [
                {'date': str(d), 'weekday': d.strftime('%A')}
                for d in expected_days
                if d not in submitted_dates
            ]

            result.append({
                'offering_id': offering.id,
                'course_name': offering.course.name,
                'section_name': offering.section.name,
                'teacher_id': offering.teacher.id,
                'teacher_name': offering.teacher.get_full_name() or offering.teacher.username,
                'teacher_email': offering.teacher.email,
                'schedule': [
                    {'day': s.day_of_week, 'day_label': dict(TeachingSchedule.DAY_CHOICES)[s.day_of_week],
                     'start_time': str(s.start_time) if s.start_time else None}
                    for s in slots
                ],
                'missed_days': missed_days,
                'total_expected': len(expected_days),
                'total_submitted': len(expected_days) - len(missed_days),
                'compliance_pct': round(
                    (len(expected_days) - len(missed_days)) / len(expected_days) * 100
                    if expected_days else 100, 1
                ),
            })

        return Response(result)


class BulkTemplateSendView(APIView):
    """POST /api/attendance/bulk-send-template/
    For each offering (optionally filtered by ?semester=, ?teacher=),
    generate the attendance template and email it to the teacher.
    Returns a summary of sent / skipped."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from .models import CourseOffering
        from .utils import send_email
        import io

        if request.user.role != 'admin':
            return Response({'error': 'Forbidden'}, status=403)

        offerings = CourseOffering.objects.select_related(
            'course', 'section', 'teacher'
        )
        if request.data.get('semester'):
            offerings = offerings.filter(section__semester_id=request.data['semester'])
        if request.data.get('teacher'):
            offerings = offerings.filter(teacher_id=request.data['teacher'])

        sent = []
        skipped = []

        for offering in offerings:
            if not offering.teacher or not offering.teacher.email:
                skipped.append({
                    'offering_id': offering.id,
                    'reason': 'No teacher or email assigned',
                })
                continue

            # Reuse the AttendanceTemplateView logic to generate the file
            try:
                from datetime import date, timedelta
                # Use the current week by default
                today = date.today()
                period = request.data.get('period', 'week')
                week_start = today - timedelta(days=today.weekday())
                if period == 'month':
                    week_end = week_start + timedelta(weeks=4, days=-1)
                else:
                    week_end = week_start + timedelta(days=4)

                fake_request = type('R', (), {
                    'query_params': {
                        'start_date': str(week_start),
                        'end_date': str(week_end),
                    }
                })()
                response = view.get(fake_request, offering.id)
                # response is an HttpResponse with xlsx content
                if not hasattr(response, 'content'):
                    skipped.append({'offering_id': offering.id, 'reason': 'Template generation failed'})
                    continue

                import base64
                attachment_b64 = base64.b64encode(response.content).decode()
                filename = f"attendance_{offering.course.name.replace(' ', '_')}_Sec{offering.section.name}_{week_start}.xlsx"

                teacher = offering.teacher
                subject = f"Attendance Template: {offering.course.name} — Sec {offering.section.name}"
                body = f"""
                <div style="font-family: Arial, sans-serif; max-width: 600px;">
                    <div style="background: #1B3A6B; padding: 16px; text-align: center;">
                        <h2 style="color: white; margin: 0;">EAU Attendance System</h2>
                    </div>
                    <div style="padding: 24px;">
                        <p>Dear {teacher.get_full_name() or teacher.username},</p>
                        <p>Please find attached the attendance template for:</p>
                        <ul>
                            <li><strong>Course:</strong> {offering.course.name}</li>
                            <li><strong>Section:</strong> {offering.section.name}</li>
                            <li><strong>Period:</strong> {week_start.strftime('%b %d')} – {week_end.strftime('%b %d, %Y')}</li>
                        </ul>
                        <p>Please fill in the attendance and return it to the admin
                        by end of the week so it can be imported into the system.</p>
                        <p style="color: #666; font-size: 12px;">
                            EAU Attendance Management System
                        </p>
                    </div>
                </div>
                """
                # Send using django's EmailMessage so we can attach binary
                from django.core.mail import EmailMessage
                from decouple import config as _config
                msg = EmailMessage(
                    subject=subject,
                    body=body,
                    from_email=_config('DEFAULT_FROM_EMAIL'),
                    to=[teacher.email],
                )
                msg.content_subtype = 'html'
                msg.attach(filename, response.content,
                           'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                msg.send(fail_silently=False)
                sent.append({
                    'offering_id': offering.id,
                    'teacher_email': teacher.email,
                    'course': offering.course.name,
                    'section': offering.section.name,
                })
            except Exception as e:
                skipped.append({'offering_id': offering.id, 'reason': str(e)})

        return Response({
            'sent_count': len(sent),
            'skipped_count': len(skipped),
            'sent': sent,
            'skipped': skipped,
        })
